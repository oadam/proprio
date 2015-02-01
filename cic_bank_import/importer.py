# -*- coding: utf-8 -*-
from bank_import.models import ImportLine
import datetime
from decimal import Decimal
from openpyxl import load_workbook

# number of rows after which we give up searching for the header
GIVE_UP_AFTER = 30
HEADER = ('Date', 'Date valeur', u'Libellé', u'Débit Euros', u'Crédit Euros',)
STOP = (None, None, None, None, None,)


class _Importer:
    def get_label(self):
        return "CIC - xlsx"

    def get_id(self):
        return "CIC-XLSX"

    def parse_row(self, row):
        date, date2, caption, debit, credit = row
        if debit is None and credit is None:
            raise ValueError(
                u'both debit and credit are empty on line {}'.format(row))
        if debit is not None and credit is not None:
            raise ValueError(
                u'both debit and credit are specified on line {}'.format(row))
        if debit is not None:
            amount = Decimal(debit).copy_negate()
        if credit is not None:
            amount = Decimal(credit)
        amount = amount.quantize(Decimal('.01'))
        return ImportLine(date=date.date(), caption=caption, amount=amount)

    def parse(self, file):
        wb = load_workbook(filename=file)
        ws = wb.active
        result = []
        read_rows = 0
        header_found = False
        for row in ws.iter_rows():
            read_rows += 1
            values = (
                row[0].value,
                row[1].value,
                row[2].value,
                row[3].value,
                row[4].value,)
            if header_found:
                if values == STOP:
                    break
                result.append(self.parse_row(values))
            else:
                if values == HEADER:
                    header_found = True
                elif read_rows > GIVE_UP_AFTER:
                    raise ValueError(u'did not find row {}'.format(HEADER))
        return result


importer = _Importer()
