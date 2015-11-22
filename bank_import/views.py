from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from forms import GenerateForm, SubmitForm
from django.http import HttpResponse, Http404, HttpResponseBadRequest
from django.utils.translation import ugettext as _, ugettext_lazy
from django.db import transaction
from main.models import Tenant, Payment
from from_settings import get_element
from whoosh.filedb.filestore import RamStorage
from whoosh.fields import TEXT, NUMERIC, Schema
from whoosh.query import Term, Or
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.writer.excel import save_virtual_workbook
import json
import re
import datetime
import itertools
from collections import defaultdict
from models import ImportedLine


IMPORTER_SETTINGS = 'PROPRIO_IMPORT_PARSERS'
MIN_SCORE = 3.0
XLSX_CONTENT_TYPE = (
    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@login_required
def forms(request):
    upload_form = GenerateForm()
    submit_form = SubmitForm()
    context = {'generate_form': upload_form, 'submit_form': submit_form}
    return render(request,  'bank_import/upload.html', context)


@login_required
def generate(request):
    if request.method != 'POST':
        raise Http404("use POST")
    form = GenerateForm(request.POST, request.FILES)
    if not form.is_valid():
        return HttpResponseBadRequest(form.errors.as_json())
    data = form.cleaned_data
    importer = get_element(IMPORTER_SETTINGS, data['type'])
    parsed_file = importer.parse(data['file'])
    parsed_file = remove_saved_lines(parsed_file)
    mapping_worbook = generate_mapping_file(parsed_file)
    today = datetime.datetime.today()
    filename = today.strftime("mapping-%m-%d_%H:%M:%S.xlsx")
    response = HttpResponse(
        mapping_worbook,
        content_type=XLSX_CONTENT_TYPE)
    cd = 'attachment; filename="{}"'.format(filename)
    response['Content-Disposition'] = cd
    return response


@login_required
def submit(request):
    if request.method != 'POST':
        raise Http404("use POST")
    form = SubmitForm(request.POST, request.FILES)
    if not form.is_valid():
        return HttpResponseBadRequest(form.errors.as_json())
    data = form.cleaned_data
    submit_mapping(data['file'])
    return HttpResponse(_("import is successful").capitalize())


class TenantPaymentMapper:
    type = 'tenant_payment'
    caption = ugettext_lazy('Tenant payment')

    def __init__(self):
        tenants = Tenant.objects.all().order_by('name')
        self.tenants = {t.id: t for t in tenants}

    def get_all_values(self):
        return self.tenants.keys()

    def get_caption(self, value):
        return self.tenants[value].name

    def get_long_caption(self, value):
        return u'{}: {}'.format(self.caption, self.get_caption(value))

    def save(self, value, line):
        tenant = self.tenants[value]
        payment = Payment(
            tenant=tenant,
            date=line.date,
            amount=line.amount,
            description=line.caption)
        payment.save()


class TenantNameGuesser:

    def __init__(self):
        tenants = Tenant.objects.all().order_by('name')
        tenant_schema = Schema(name=TEXT(stored=True), id=NUMERIC(stored=True))
        tenant_storage = RamStorage()
        tenant_ix = tenant_storage.create_index(tenant_schema)
        tenant_writer = tenant_ix.writer()
        for t in tenants:
            tenant_writer.add_document(id=t.id, name=t.name.lower())
        tenant_writer.commit()
        self.index = tenant_ix

    def guess(self, line):
        with self.index.searcher() as searcher:
            words = re.split('\W+', line.caption)
            query = Or([Term("name", t.lower()) for t in words])
            result = searcher.search(query)
            matches = [
                (('tenant_payment', r['id'],), r.score,)
                for r in result]
            return matches


mapper_factories = (lambda: TenantPaymentMapper(),)

guesser_factories = (lambda: TenantNameGuesser(),)


def value_to_combo_entry(mapper, value):
    id = json.dumps((mapper.type, value,))
    caption = mapper.get_long_caption(value)
    return (id, caption,)


def guess(guessers, mappers, line):
    guesses_map = defaultdict(int)
    for g in guessers:
        guess = g.guess(line)
        for value, score in guess:
            guesses_map[value] += score
    guesses = sorted(guesses_map.items(), key=lambda g: -g[1])
    if not guesses:
        return None
    else:
        (mapper_type, value), score = guesses[0]
        if score < MIN_SCORE:
            return None
        else:
            mapper = mappers[mapper_type]
            return value_to_combo_entry(mapper, value)


def remove_saved_lines(lines):
    saved = ImportedLine.objects.all()
    all_ids = set([(l.date, l.amount, l.caption,) for l in saved])
    return [l for l in lines
            if (l.date, l.amount, l.caption,) not in all_ids]


def fill_all_mappings(worksheet, mappers):
    hardcoded_choices = [
        ('', _('Decide later')),
        ('HIDE', _('Hide line definitively')),
    ]
    mapper_choices = [
        value_to_combo_entry(m, v)
        for m in mappers
        for v in m.get_all_values()]
    current_row = 1
    caption_occurences = defaultdict(int)
    for id, caption in itertools.chain(hardcoded_choices, mapper_choices):
        caption_occurences[caption] += 1
        occurences = caption_occurences[caption]
        if occurences > 1:
            caption += '_{}'.format(occurences)
        worksheet.cell(column=2, row=current_row).value = id
        worksheet.cell(column=1, row=current_row).value = caption
        current_row += 1


def parse_caption_to_id(all_mapping_worksheet):
    result = {}
    for row in range(1, all_mapping_worksheet.max_row + 1):
        caption = all_mapping_worksheet.cell(row=row, column=1).value
        if caption is None:
            # skip empty rows (max_row is not very reliable)
            continue
        id = all_mapping_worksheet.cell(row=row, column=2).value
        # for Decide later mapping
        if id is None:
            id = ''
        result[caption] = id
    return result


def get_mappers_and_guessers():
    mappers = [m() for m in mapper_factories]
    mappers_map = {m.type: m for m in mappers}
    guessers = [g() for g in guesser_factories]
    return (mappers_map, guessers,)


def generate_mapping_file(lines):
    wb = Workbook()
    main_sheet = wb.active
    main_sheet.title = _('mapping')
    mapping_sheet = wb.create_sheet()
    mapping_sheet.title = _('possible_mappings')

    mappers, guessers = get_mappers_and_guessers()

    fill_all_mappings(mapping_sheet, mappers.values())
    caption_to_id = parse_caption_to_id(mapping_sheet)
    id_to_caption = dict(reversed(item) for item in caption_to_id.items())

    wb.create_named_range('all_captions', mapping_sheet, 'A1:A1048576')
    dv = DataValidation(
        type="list",
        formula1='all_captions',
        allow_blank=True,
        showDropDown=True)
    dv.ranges.append('D1:D1048576')
    main_sheet.add_data_validation(dv)

    main_sheet['A1'] = _('date').capitalize()
    main_sheet['B1'] = _('amount').capitalize()
    main_sheet['C1'] = _('caption').capitalize()
    main_sheet['D1'] = _('mapping').capitalize()
    for i in range(len(lines)):
        line = lines[i]
        best_guess = guess(guessers, mappers, line)
        if best_guess is None:
            best_id = ''  # decide later
        else:
            best_id, useless_caption = best_guess
        rownum = i+2
        main_sheet.cell(column=1, row=rownum).value = line.date
        main_sheet.cell(column=2, row=rownum).value = line.amount
        main_sheet.cell(column=3, row=rownum).value = line.caption
        main_sheet.cell(column=4, row=rownum).value = id_to_caption[best_id]

    return save_virtual_workbook(wb)


@transaction.atomic
def submit_lines(lines):
    mapper_map, _guessers = get_mappers_and_guessers()
    for line in lines:
        mapping = line.mapping
        # skip non-mapped lines
        if mapping == '':
            continue
        # save the mapping to avoid reimporting next time
        line.save()
        # if this is a simple hide there is nothing more to do
        if mapping == 'HIDE':
            continue
        mapper_type, value = json.loads(mapping)
        mapper = mapper_map[mapper_type]
        mapper.save(value, line)


def submit_mapping(file):
    wb = load_workbook(file, guess_types=True)
    main_sheet = wb[_('mapping')]
    mapping_sheet = wb[_('possible_mappings')]
    caption_to_id = parse_caption_to_id(mapping_sheet)
    result = []
    for row in range(2, main_sheet.max_row + 1):
        date = main_sheet.cell(column=1, row=row).value
        if date is None:
            # skip empty rows (max_row is not very reliable)
            continue
        amount = main_sheet.cell(column=2, row=row).value
        caption = main_sheet.cell(column=3, row=row).value
        mapping_caption = main_sheet.cell(column=4, row=row).value
        if mapping_caption not in caption_to_id:
            raise ValueError(_('unknow mapping').format(mapping_caption))
        mapping = caption_to_id[mapping_caption]
        result.append(ImportedLine(
            date=date,
            amount=amount,
            caption=caption,
            mapping=mapping))
    submit_lines(result)
